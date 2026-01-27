"""
Report Generator Module

Merges and processes event data into a consolidated Time Trials report.
"""

import csv
import json
import os
from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ReportGenerator:
    """Generates consolidated Time Trials reports from MSR data."""

    def __init__(self, export_dir: str):
        """
        Initialize the report generator.

        Args:
            export_dir: Directory containing exported CSV files
        """
        self.export_dir = export_dir
        self.entrylist_file = os.path.join(export_dir, "entrylist.csv")
        self.attendees_file = os.path.join(export_dir, "attendees.csv")
        self.assignments_file = os.path.join(export_dir, "assignments.csv")
        self.assignments_json_file = os.path.join(export_dir, "assignments.json")

    def _read_csv(self, filepath: str) -> List[Dict]:
        """Read a CSV file into a list of dictionaries."""
        rows = []
        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append(row)
        return rows

    def _read_json(self, filepath: str) -> Any:
        """Read a JSON file."""
        with open(filepath, "r", encoding="utf-8") as f:
            return json.load(f)

    def _get_driver_key(self, row: Dict) -> str:
        """Create a unique key for a driver based on first and last name."""
        first = (row.get("firstName") or "").strip().lower()
        last = (row.get("lastName") or "").strip().lower()
        return f"{first}|{last}"

    def _parse_segment(self, segment: str) -> Optional[str]:
        """Extract the day from segment (Friday/Saturday/Sunday)."""
        if not segment:
            return None
        segment_lower = segment.lower()
        if "friday" in segment_lower:
            return "Friday"
        elif "saturday" in segment_lower:
            return "Saturday"
        elif "sunday" in segment_lower:
            return "Sunday"
        return None

    def _is_time_trials(self, group: str) -> bool:
        """Check if this is a Time Trials entry."""
        if not group:
            return False
        return "time trials" in group.lower()

    def _is_instructor(self, group: str) -> bool:
        """Check if this is an Instructing entry."""
        if not group:
            return False
        return "instructing" in group.lower()

    def _is_advanced_hpde(self, group: str) -> bool:
        """Check if this is an Advanced HPDE entry."""
        if not group:
            return False
        return "advanced hpde" in group.lower()

    def _is_worker_only(self, segment: str) -> bool:
        """Check if this is a worker-only entry (not a track event)."""
        if not segment:
            return True
        return "workers" in segment.lower()

    def _get_participation_type(self, is_instructor: bool, is_ayce: bool) -> str:
        """
        Determine participation type for pivot table analysis.

        Returns one of:
        - TT + Instructor + AYCE
        - TT + Instructor
        - TT + AYCE
        - TT Only
        """
        if is_instructor and is_ayce:
            return "TT + Instructor + AYCE"
        elif is_instructor:
            return "TT + Instructor"
        elif is_ayce:
            return "TT + AYCE"
        else:
            return "TT Only"

    def _get_day_count(self, day_count: int) -> str:
        """
        Determine day count category for pivot table analysis.

        Returns: '1 Day', '2 Days', or '3 Days'
        """
        if day_count == 1:
            return "1 Day"
        elif day_count == 2:
            return "2 Days"
        elif day_count >= 3:
            return "3 Days"
        else:
            return ""

    def _format_days_string(self, days_tt: Set[str]) -> Tuple[str, int]:
        """
        Format days participation into display string and count.

        Returns:
            Tuple of (formatted days string, day count)
        """
        has_fri = "Friday" in days_tt
        has_sat = "Saturday" in days_tt
        has_sun = "Sunday" in days_tt
        day_count = sum([has_fri, has_sat, has_sun])

        if day_count == 3:
            return "All 3", day_count
        if day_count == 2:
            if has_fri and has_sat:
                return "Fri/Sat", day_count
            if has_fri and has_sun:
                return "Fri/Sun", day_count
            return "Sat/Sun", day_count
        if day_count == 1:
            if has_fri:
                return "Friday", day_count
            if has_sat:
                return "Saturday", day_count
            return "Sunday", day_count
        return "", day_count

    def _format_vehicle_string(self, driver: Dict) -> str:
        """Combine year, make, model into single vehicle string."""
        parts = []
        if driver["year"]:
            parts.append(str(driver["year"]))
        if driver["make"]:
            parts.append(driver["make"])
        if driver["model"]:
            parts.append(driver["model"])
        return " ".join(parts)

    def _write_driver_row(self, ws, row_num: int, driver: Dict, thin_border, center_cols: List[int]) -> None:
        """Write a single driver's data row to the worksheet."""
        # Format days and get count
        days_str, day_count = self._format_days_string(driver["days_tt"])

        # Calculate derived values
        is_ayce = driver["is_tt"] and driver["is_advanced_hpde"]
        class_group = self._get_class_group(driver["class"])
        day_count_str = self._get_day_count(day_count)
        participation_type = self._get_participation_type(driver["is_instructor"], is_ayce)
        vehicle = self._format_vehicle_string(driver)

        # Build row data
        row_data = [
            driver["firstName"],
            driver["lastName"],
            driver["email"],
            driver["memberId"],
            driver["class"],
            class_group,
            driver["vehicleNumber"],
            vehicle,
            driver["color"],
            driver["tireBrand"],
            driver["sponsor"],
            days_str,
            day_count_str,
            "Yes" if driver["is_instructor"] else "No",
            "Yes" if is_ayce else "No",
            participation_type,
            driver["status"],
        ]

        # Write cells with formatting
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if col in center_cols:
                cell.alignment = Alignment(horizontal="center")

    def _build_tire_lookup(self) -> Dict[str, str]:
        """Build lookup dictionary for tire brands from assignments JSON."""
        tire_lookup = {}
        if not os.path.exists(self.assignments_json_file):
            return tire_lookup

        assignments_data = self._read_json(self.assignments_json_file)
        for assignment in assignments_data.get("assignments", []):
            key = self._get_driver_key(assignment)
            group = assignment.get("group", "")
            tire = assignment.get("tireBrand", "")
            # Only capture tire from Time Trials entries
            if key and key != "|" and self._is_time_trials(group) and tire:
                tire_lookup[key] = tire

        return tire_lookup

    def _build_attendee_lookup(self, attendees: List[Dict]) -> Dict[str, Dict]:
        """Build lookup dictionary for attendee data."""
        attendee_lookup = {}
        for att in attendees:
            key = self._get_driver_key(att)
            if key and key != "|":
                attendee_lookup[key] = att
        return attendee_lookup

    def _initialize_driver_record(self, entry: Dict) -> Dict:
        """Initialize a new driver record with default values."""
        return {
            "firstName": entry.get("firstName", "").strip(),
            "lastName": entry.get("lastName", "").strip(),
            "class": "",
            "make": "",
            "model": "",
            "year": "",
            "vehicleNumber": "",
            "color": "",
            "sponsor": "",
            "tireBrand": "",
            "is_tt": False,
            "is_instructor": False,
            "is_advanced_hpde": False,
            "days_tt": set(),
            "days_instructor": set(),
            "days_advanced": set(),
            "email": "",
            "memberId": "",
            "status": "",
        }

    def _update_driver_with_tt_data(self, driver: Dict, entry: Dict, day: Optional[str]) -> None:
        """Update driver record with Time Trials data."""
        driver["is_tt"] = True
        if day:
            driver["days_tt"].add(day)
        # Capture vehicle info from TT entry
        if entry.get("class"):
            driver["class"] = entry.get("class", "")
        if entry.get("make"):
            driver["make"] = entry.get("make", "")
        if entry.get("model"):
            driver["model"] = entry.get("model", "")
        if entry.get("year"):
            driver["year"] = entry.get("year", "")
        if entry.get("vehicleNumber"):
            driver["vehicleNumber"] = entry.get("vehicleNumber", "")
        if entry.get("color"):
            driver["color"] = entry.get("color", "")
        if entry.get("sponsor"):
            driver["sponsor"] = entry.get("sponsor", "")

    def _process_entry(self, entry: Dict, drivers: Dict[str, Dict]) -> None:
        """Process a single entry and update driver records."""
        # Skip worker-only entries
        if self._is_worker_only(entry.get("segment", "")):
            return

        driver_key = self._get_driver_key(entry)
        if not driver_key or driver_key == "|":
            return

        group = entry.get("group", "")
        segment = entry.get("segment", "")
        day = self._parse_segment(segment)

        # Initialize driver record if new
        if driver_key not in drivers:
            drivers[driver_key] = self._initialize_driver_record(entry)

        driver = drivers[driver_key]

        # Track participation types
        if self._is_time_trials(group):
            self._update_driver_with_tt_data(driver, entry, day)

        if self._is_instructor(group):
            driver["is_instructor"] = True
            if day:
                driver["days_instructor"].add(day)

        if self._is_advanced_hpde(group):
            driver["is_advanced_hpde"] = True
            if day:
                driver["days_advanced"].add(day)

    def _enrich_drivers_with_metadata(
        self, drivers: Dict[str, Dict], attendee_lookup: Dict, tire_lookup: Dict
    ) -> None:
        """Add attendee info and tire data to driver records."""
        for driver_key, driver in drivers.items():
            if driver_key in attendee_lookup:
                att = attendee_lookup[driver_key]
                driver["email"] = att.get("email", "")
                driver["memberId"] = att.get("memberId", "")
                driver["status"] = att.get("status", "")
            if driver_key in tire_lookup:
                driver["tireBrand"] = tire_lookup[driver_key]

    def _create_workbook_with_headers(self) -> Tuple:
        """Create Excel workbook with formatted headers."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Time Trials Report"

        headers = [
            "First Name",
            "Last Name",
            "Email",
            "Member ID",
            "Class",
            "Class Group",
            "Vehicle #",
            "Vehicle",
            "Color",
            "Tire",
            "Sponsor",
            "Days (TT)",
            "Day Count",
            "Instructor",
            "AYCE",
            "Participation Type",
            "Status",
        ]

        # Style settings
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        return wb, ws, headers, thin_border

    def _auto_adjust_column_widths(self, ws, headers: List[str], num_rows: int) -> None:
        """Auto-adjust column widths based on content."""
        for col in range(1, len(headers) + 1):
            max_length = len(headers[col - 1])
            for row in range(2, num_rows + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col)].width = adjusted_width

    def _get_class_group(self, tt_class: str) -> str:
        """
        Determine class grouping for pivot table analysis.

        Groups:
        - Max (Max 1, Max 2, Max 5, Max 6, etc.)
        - Sport (Sport 1, Sport 2, Sport 3, Sport 4, etc.)
        - Tuner (Tuner 1, Tuner 2, Tuner 3, Tuner 4, etc.)
        - Unlimited (Unlimited 1, Unlimited 2, etc.)
        - Other (anything else or empty)
        """
        if not tt_class:
            return "Other"

        class_lower = tt_class.lower().strip()

        if class_lower.startswith("max"):
            return "Max"
        if class_lower.startswith("sport"):
            return "Sport"
        if class_lower.startswith("tuner"):
            return "Tuner"
        if class_lower.startswith("unlimited"):
            return "Unlimited"
        return "Other"

    def generate_tt_report(self, output_path: Optional[str] = None) -> Tuple[str, int]:
        """
        Generate consolidated Time Trials report.

        Only includes drivers who participated in Time Trials.
        Tracks if they also instructed and/or did Advanced HPDE (AYCE).

        Args:
            output_path: Path for output file (default: same dir as input)

        Returns:
            Tuple of (path to generated report file, number of drivers)
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. Install with: pip install openpyxl"
            )

        # Read and prepare data
        entrylist = self._read_csv(self.entrylist_file)
        attendees = self._read_csv(self.attendees_file)
        tire_lookup = self._build_tire_lookup()
        attendee_lookup = self._build_attendee_lookup(attendees)

        # Process entries - group by driver
        drivers: Dict[str, Dict] = {}
        for entry in entrylist:
            self._process_entry(entry, drivers)

        # Filter to only Time Trials participants
        tt_drivers = {k: v for k, v in drivers.items() if v["is_tt"]}

        # Add attendee info and tire data
        self._enrich_drivers_with_metadata(tt_drivers, attendee_lookup, tire_lookup)

        # Create Excel workbook with headers
        wb, ws, headers, thin_border = self._create_workbook_with_headers()

        # Sort drivers and write data rows
        sorted_drivers = sorted(
            tt_drivers.values(), key=lambda d: (d["lastName"].lower(), d["firstName"].lower())
        )

        center_cols = [6, 12, 13, 14, 15, 16]
        for row_num, driver in enumerate(sorted_drivers, 2):
            self._write_driver_row(ws, row_num, driver, thin_border, center_cols)

        # Format worksheet
        self._auto_adjust_column_widths(ws, headers, len(sorted_drivers))
        ws.freeze_panes = "A2"

        # Set output path
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(self.export_dir, f"tt_report_{timestamp}.xlsx")

        wb.save(output_path)

        return output_path, len(sorted_drivers)


def generate_report(
    export_dir: str, output_path: Optional[str] = None, verbose: bool = False
) -> str:
    """
    Generate Time Trials report from exported data.

    Args:
        export_dir: Directory containing exported CSV files
        output_path: Optional output file path
        verbose: Print progress messages

    Returns:
        Path to generated report
    """
    generator = ReportGenerator(export_dir)

    if verbose:
        print("Generating Time Trials report...")

    report_path, driver_count = generator.generate_tt_report(output_path)

    if verbose:
        print(f"  [OK] Report generated: {report_path}")
        print(f"       {driver_count} Time Trials drivers included")

    return report_path
